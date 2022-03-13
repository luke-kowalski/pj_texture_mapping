from openpyxl import load_workbook

wb = load_workbook(filename="pj.xlsx")
wb2 = load_workbook(filename="grafiki.xlsx")

ws = wb.active
ws2 = wb2.active


def get_data_from_excel():

    with open("texture_mapping.xml", "w", encoding="UTF-8", errors="ignore") as f:

        f.write(
            '<?xml version="1.0" encoding="UTF-8" standalone="yes" ?>\n\n<VARIABLES_MAPPING>\n'
        )

        for sheet in wb.sheetnames:
            ws_s = wb[sheet]
            for col in ws_s["A"]:
                f.write(f'\n<VARIABLE nav_name="{col.value}" imos_name="{sheet}"/>\n\n')
                for row in range(1, ws2.max_row + 1):
                    a = ws2["A" + str(row)].value
                    b = ws2["B" + str(row)].value
                    f.write(f'\t<VALUE imos_value="{a}" nav_value="{b}"/>\n')
                f.write("\n</VARIABLE>\n")

        f.write("\n</VARIABLES_MAPPING>")


get_data_from_excel()

# for sheet in wb.sheetnames:
#     ws = wb[sheet]
#     for col in ws['A']:
#         print(f'<VARIABLE nav_name="{col.value}" imos_name="{sheet}"/>')

# from openpyxl import Workbook, load_workbook

# workbook = load_workbook(filename="pj.xlsx")

# ws = workbook['Arkusz1']
# ws2 = Workbook()

# for col in ws['A']:
#     for col2 in ws['B']:
#         if col2.value:
#             print(f"{col.value}_{col2.value}")
#     ws2.append
